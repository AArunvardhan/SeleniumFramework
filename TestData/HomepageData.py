import openpyxl


class HomePageData:
    test_data = [{"Firstname":"Arun","Email":"vardhan@gmail.com","Gender":"Male","Password":9876543},
                 {"Firstname":"varun","Email":"var@gmail.com","Gender":"Male","Password":123456}]
    @staticmethod
    def get_test_data(test_case_name):
        dict1 = {}
        book = openpyxl.load_workbook("D:/PytestDataBook.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [dict1]